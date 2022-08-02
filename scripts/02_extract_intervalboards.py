import datetime
import glob
import os
import shutil
import numpy as np
import pandas as pd
from csv import writer
# import textwrap
from pathlib import Path
import utm
from openpyxl import load_workbook
# import matplotlib.pyplot as plt
# import seaborn as sns
# from dataclasses import dataclass

'''
Add text here about the script.
'''


def readIntervalBoard(path_in, filename, version, path_out, fname_newSnow):

    #put in function
    wb = load_workbook(filename)
    ws = wb.active

    # metadata
    site = ws['D4'].value
    location = ws['D5'].value
    site_id = ws['D6'].value
    observer = ws['D7'].value
    date = ws['C8'].value
    time = ws['E8'].value #.strptime("%H:%M") --> either correct for 'n/o' or leave as string
    lat = ws['C9'].value
    lon = ws['E9'].value

    #data
    hgtA = ws['D12'].value # sample A
    hgtB = ws['D13'].value # sample B
    hgtC = ws['D14'].value # sample C

    sweA = ws['E12'].value
    sweB = ws['E13'].value
    sweC = ws['E14'].value

    denA = ws['F12'].value
    denB = ws['F13'].value
    denC = ws['F14'].value

    melt = ws['E15'].value
    comments = ws['B17'].value

    # conditional statements to convert strings > np.nan, and round density

    # convert str to np.nan
    if isinstance(hgtA, str): # height A
        hgtA = np.nan

    if isinstance(hgtB, str): # height B
        hgtB = np.nan

    if isinstance(hgtC, str): # height C
        hgtC = np.nan


    if isinstance(sweA, str): # swe A
        sweA = np.nan

    if isinstance(sweB, str): # swe B
        sweB = np.nan

    if isinstance(sweC, str): # swe C
        sweC = np.nan


    # round density
    if isinstance(denA, float) or isinstance(denA, int): # density A
        denA = int(np.ceil(denA))
    else:
        denA = np.nan

    if isinstance(denB, float) or isinstance(denB, int): # density B
        denB = int(np.ceil(denB))
    else:
        denB = np.nan

    if isinstance(denC, float) or isinstance(denC, int): # density C
        denC = int(np.ceil(denC))
    else:
        denC = np.nan


    # create dictionary to store data
    dict = {}
    dict = {'Latitude': round(lat, 5),
            'Longitude': round(lon, 5),
            'Date': date,
            'Local \nTime': time,
            'Week No.': date.isocalendar()[1]-1, #returns tuple, want 2nd item
            'PitID': site_id.split('_')[0],
            'State': site_id[:2],
            'Location': location,
            'Site': site,
            'HN (cm) A': hgtA,
            'HN (cm) B': hgtB,
            'HN (cm) C': hgtC,
            # 'AVG HN (cm)': avgH,
            'SWE (mm) A': sweA,
            'SWE (mm) B': sweB,
            'SWE (mm) C': sweC,
            # 'AVG SWE (mm)': avgS,
            'DEN (kg/m^3) A': denA,
            'DEN (kg/m^3) B': denB,
            'DEN (kg/m^3) C': denC,
            # 'AVG DEN (kg/m^3)': avgD,
            'Melt': melt,
            'Comments': comments,}

    dict.update(dict)

    rows_list.append(dict)


#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
if __name__ == "__main__":

    # static variables
    version = 'v01' # working version
    hsphere = 'N' # northern hemisphere
    year_to_process = 2021

    # paths
    path_in = Path('/Users/mamason6/Documents/snowex/campaigns/TS-21/3_interval-board-adjusted/')
    path_out = Path('/Users/mamason6/Documents/snowex/core-datasets/ground/interval-boards/output/')

    # create output filename
    fname_newSnow = path_out.joinpath('SNEX21_TS_IB_Summary_newSnow_' + version +'.csv')
    fname_newSnow_avgs = path_out.joinpath('SNEX21_TS_IB_Summary_newSnow_averages_' + version +'.csv')
    fname_newSnow_sites_avg = path_out.joinpath('SNEX21_TS_IB_Summary_newSnow_sites_avg_' + version +'.csv')
    fname_newSnow_locations_avg = path_out.joinpath('SNEX21_TS_IB_Summary_newSnow_locations_avg_' + version +'.csv')

    # empty dataframe
    df = pd.DataFrame()

    # empty rows list
    rows_list = []


    # loop over all interval board sheets
    for filename in sorted(path_in.glob('*.xlsx')):
    # for filename in sorted(path_in.glob('*COF*')):
    # for filename in sorted(path_in.glob('*COFEB1_20210127_INTERVAL_BOARD.xlsx')):

        print('running file .../', filename.name)

        r = readIntervalBoard(path_in, filename, version, path_out, fname_newSnow)

    df = pd.DataFrame(rows_list)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# compute average on dataframe

    # height of new snow (HN (cm))
    df['avg HN (cm)'] = df[['HN (cm) A', 'HN (cm) B', 'HN (cm) C']].mean(axis=1)
    df['avg HN (cm)']= df['avg HN (cm)'].round(decimals=1)

    # Snow water equivalent (SWE (mm))
    df['avg SWE (mm)'] = df[['SWE (mm) A', 'SWE (mm) B', 'SWE (mm) C']].mean(axis=1)
    df['avg SWE (mm)']= df['avg SWE (mm)'].round(decimals=0) #'DEN (kg/m^3) A'

    # df['avg DEN (kg/m^3)'] = df[['DEN (kg/m^3) A', 'DEN (kg/m^3) B', 'DEN (kg/m^3) C']].mean(axis=1)
    # df['avg DEN (kg/m^3)']= df['avg DEN (kg/m^3)'].apply(np.ceil) # "raw density"

    df['avg DEN (kg/m^3)'] = (df['avg SWE (mm)'] / df['avg HN (cm)']) * 100
    df['avg DEN (kg/m^3)']= df['avg DEN (kg/m^3)'].apply(np.ceil)

    # create dataframe for just the averages
    df_avgs = df[['Latitude', 'Longitude', 'Date', 'Local \nTime', 'Week No.', 'PitID',
           'State', 'Location', 'Site', 'avg HN (cm)', 'avg SWE (mm)', 'avg DEN (kg/m^3)',
            'Melt', 'Comments']]

    # remove any with week number >11
    df_avgs = df_avgs[df_avgs['Week No.'] <= 11] #11 flight weeks for 2021

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# averaging location and site locations (groupby).mean()
    avg_site_df = df.groupby('Site')[['Latitude', 'Longitude']].mean().reset_index()
    avg_loc_df = df.groupby('Location')[['Latitude', 'Longitude']].mean().reset_index()
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# saving outputs

    # convert to csv and save
    df.to_csv(fname_newSnow, sep=',', index=False, na_rep=' ')
    df_avgs.to_csv(fname_newSnow_avgs, sep=',', index=False, na_rep=' ')
    avg_site_df.to_csv(fname_newSnow_sites_avg, sep=',', index=False, na_rep=' ')
    avg_loc_df.to_csv(fname_newSnow_locations_avg, sep=',', index=False, na_rep=' ')

print('..... Script is Complete !  .....')



# #https://realpython.com/openpyxl-excel-spreadsheets-python/
#         @dataclass
#         class Metadata:
#             interval_board_name:str
#             location:str
#             site_id:str
#             observer:str
#             date:datetime
#             time: datetime
#             latitude:float
#             longitude:float
#             comments:
#
#         @dataclass
#         class Data:
#             depth:
