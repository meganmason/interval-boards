import datetime
import glob
import os
import shutil
import numpy as np
import pandas as pd
from csv import writer
from pathlib import Path
import utm
from openpyxl import load_workbook
# import matplotlib.pyplot as plt
# import seaborn as sns
# from dataclasses import dataclass

'''
Add text here about the script.
'''


def readIntervalBoard(path_in, filename, version, path_out, fname_newSnow):

    #put in function
    wb = load_workbook(filename, read_only=True)#, data_only=True)
    ws = wb.active

    # metadata
    site = ws['D4'].value
    location = ws['D5'].value
    site_id = ws['D6'].value
    observer = ws['D7'].value
    date = ws['C8'].value
    time = ws['E8'].value #.strptime("%H:%M") --> either correct for 'n/o' or leave as string
    lat = ws['C9'].value
    lon = ws['E9'].value

    #data
    hgtA = ws['D12'].value # sample A
    hgtB = ws['D13'].value # sample B
    hgtC = ws['D14'].value # sample C

    sweA = ws['E12'].value
    sweB = ws['E13'].value
    sweC = ws['E14'].value

    denA = ws['F12'].value
    denB = ws['F13'].value
    denC = ws['F14'].value

    melt = ws['E15'].value
    comments = ws['B17'].value


    # compute averages

    if hgtA == 'n/a':
        avgH = None
        avgS = None
        avgD = None

    else:
        ms_raw = [hgtA, hgtB, hgtC, sweA, sweB, sweC, denA, denB, denC]
        ms_values = [np.nan if x == 'n/o' else x for x in ms_raw]

        avgH = np.nanmean(ms_values[0:3])
        avgS = np.nanmean(ms_values[3:6])
        avgD = np.nanmean(ms_values[6:9])



    # create dictionary to store data
    dict = {}
    dict = {'Latitude': lat,
            'Longitude': lon,
            'Date': date,
            'Local \nTime': time,
            'Week No.': date.isocalendar()[1]-1, #returns tuple, want 2nd item
            'PitID': site_id.split('_')[0],
            'State': site_id[:2],
            'Location': location,
            'Site': site,
            # 'HN (cm) A': hgtA,
            # 'HN (cm) B': hgtB,
            # 'HN (cm) C': hgtC,
            'AVG HN (cm)': avgH,
            # 'SWE (mm) A': sweA,
            # 'SWE (mm) B': sweB,
            # 'SWE (mm) C': sweC,
            'AVG SWE (mm)': avgS,
            # 'DEN (kg/m^3) A': denA,
            # 'DEN (kg/m^3) B': denB,
            # 'DEN (kg/m^3) C': denC,
            'AVG DEN (kg/m^3)': avgD,
            'Melt': melt,
            'Comments': comments,}

    dict.update(dict)

    rows_list.append(dict)


#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
if __name__ == "__main__":

    # static variables
    version = 'v01' # working version
    hsphere = 'N' # northern hemisphere
    year_to_process = 2021

    # paths
    path_in = Path('/Users/mamason6/Documents/snowex/core-datasets/ground/interval-boards/output/clean_xls')
    path_out = Path('/Users/mamason6/Documents/snowex/core-datasets/ground/interval-boards/output/')

    # create output filename
    fname_newSnow = path_out.joinpath('SNEX21_TS_IB_Summary_newSnow_avgs' + version +'.csv')

    # empty dataframe
    df = pd.DataFrame()

    # empty rows list
    rows_list = []


    # loop over all interval board sheets
    for filename in path_in.glob('*.xlsx'):
        # if 'COCP' in filename.name:
        print('running file .../', filename.name)

        r = readIntervalBoard(path_in, filename, version, path_out, fname_newSnow)

    df = pd.DataFrame(rows_list)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# saving outputs

    # convert to csv and save
    df.to_csv(fname_newSnow, sep=',', index=False, na_rep=' ')


print('..... Script is Complete !  .....')



# #https://realpython.com/openpyxl-excel-spreadsheets-python/
#         @dataclass
#         class Metadata:
#             interval_board_name:str
#             location:str
#             site_id:str
#             observer:str
#             date:datetime
#             time: datetime
#             latitude:float
#             longitude:float
#             comments:
#
#         @dataclass
#         class Data:
#             depth:
