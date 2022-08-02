import datetime
import glob
from pathlib import Path
import shutil
import utm
from openpyxl import load_workbook

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# SET UP

# paths
path_in  = Path('/Users/mamason6/Documents/snowex/campaigns/TS-21/2_interval-board-cleaned/')# input path
# path_in  = Path('/Users/mamason6/Documents/snowex/campaigns/TS-21/3_interval-board-site-coords-adjusted/')# input path
# path_out = Path('/Users/mamason6/Documents/snowex/core-datasets/ground/interval-boards/output/clean_xls/') # output path
# path_out.mkdir(parents=False, exist_ok=True)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# FORMATING CORRECTIONS

#         # 1). add colon in time (HH:MM)
#
#         '''
#         NOTE:
#             This was a one time run to clear up 'time' entries
#             if run again the times will NO longer be 'floats'
#         '''
#
# # loop files and copy to output directory
# for filename in path_in.rglob('*INTERVAL_BOARD*'):
# # for filename in path_in.rglob('*COSBBR_20210210_INTERVAL_BOARD.xlsx*'):
#     print(filename.name)
#
#     # load workbook
#     wb = load_workbook(filename)
#     ws = wb.active
#
#     # grab time value
#     time = ws['E8'].value
#     print(filename.name, time)
#
#     if isinstance(time, float): # if time is a float (i.e. 1355.0, no colon was entered)
#         time = str(time) # convert to string
#         if len(time)<6: # AM hours before 10am
#             ws['E8'].value = time[:1] + ':' + time[1:3]
#
#         else: # hours after 10
#             ws['E8'].value = time[:2] + ':' + time[2:4]
#
#
#         wb.save(filename)
#
#         print('the formatted time value is now:', ws['E8'].value)


#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        #2). make time a datetime type ----- plan B --> string type :/

'''
# NOTE:
    This was a one time run to convert all time strings
    back to datetime type
'''

# for filename in path_in.rglob('*INTERVAL_BOARD*'):
#
#     # load workbook
#     wb = load_workbook(filename)
#     ws = wb.active
#
#     # grab time value
#     t = ws['E8'].value
#
#     if isinstance(t, datetime.time):
#         ws['E8'].value = t.strftime("%H:%M") #creates a string
#
#     if isinstance(t, str):
#         if t.lower() == 'n/o':
#             ws['E8'].value = 'n/o'
#
#         else:
#             ws['E8'].value = t
#
#     # print(type(ws['E8'].value))
#
#     wb.save(filename)


#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        #3). standardize Mores creek summit NR and NA to n/o
#
# def ListOfMeasurements(filename):
#
#     # load workbook
#     wb = load_workbook(filename)
#     ws = wb.active
#
#     hgtA = ws['D12'].value # sample A
#     hgtB = ws['D13'].value # sample B
#     hgtC = ws['D14'].value # sample C
#
#     sweA = ws['E12'].value
#     sweB = ws['E13'].value
#     sweC = ws['E14'].value
#
#     denA = ws['F12'].value
#     denB = ws['F13'].value
#     denC = ws['F14'].value
#
#     ms = [hgtA, hgtB, hgtC, sweA, sweB, sweC, denA, denB, denC]
#
#     return ms
#
# for filename in path_in.rglob('*undefined*'):
#
#     print(filename.name)
#     ms = ListOfMeasurements(filename) # measurement list
#     ms1 = ['n/o' if x == 'NR' else x for x in ms] #corrected measurement list, NR
#     ms2 = ['n/o' if x == 'NA' else x for x in ms1] # corrected measurement list, NA
#
#     ws['D12'].value = ms2[0]
#     ws['D13'].value = ms2[1]
#     ws['D14'].value = ms2[2]
#     ws['E12'].value = ms2[3]
#     ws['E13'].value = ms2[4]
#     ws['E14'].value = ms2[5]
#     ws['F12'].value = ms2[6]
#     ws['F13'].value = ms2[7]
#     ws['F14'].value = ms2[8]
#
#     wb.save(filename)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        #4). convert Fraser's Eastings/Northings to lat/lon

# zone = 13
#
# for filename in sorted(path_in.rglob('*COFS*')):
#     print(filename.name)
#
#     # load workbook
#     wb = load_workbook(filename)
#     ws = wb.active
#
#     # get easting/northing
#     e = ws['C9'].value
#     n = ws['E9'].value
#
#     if e > 90:
#
#         # convert
#         lat, lon = utm.to_latlon(e, n, zone, "Northern")
#         print('lat: ', lat)
#         print('lon: ', lon)
#
#         # replace in xls cell
#         ws['C9'].value = lat.round(5)
#         ws['E9'].value = lon.round(5)
#
#         wb.save(filename)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        #5). fix Mores Creek Summit "undefined" issue

# for filename in sorted(path_in.rglob('*IDBRM2*')):
#     # print(filename.name)
#
#     site_code = filename.name.split('_')[0]
#
#     # load workbook
#     wb = load_workbook(filename)
#     ws = wb.active
#
#     dt = ws['D6'].value.split("_")[-1]
#     new_siteid = site_code + '_' + dt
#     print(new_siteid)
#
#     ws['D6'].value = new_siteid
#
#     wb.save(filename)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        #6). update coordinates (should only be mores creek, but program for all)
outpath = Path('/Users/mamason6/Documents/snowex/campaigns/TS-21/3_interval-board-adjusted')

# dictionary of names (adjusted from collection period) and some coords corrected
coord_dict= {
"COCPCP":	["Cameron Peak",	40.56390,	-105.86730],
"COCPMR":	["Michigan River",	40.51860,	-105.89160],
"COFEB1":	["B1",	            39.90696,	-105.87789],
"COFEB2":	["B2",	            39.90702,	-105.87905],
"COFEJF":	["JF",	            39.90566,	-105.88304],
"COFEJS":	["JS",	            39.90566,	-105.88305],
"COFS01":	["FS1",	            39.92084,	-105.87384],
"COFS02":	["FS2",	            39.92416,	-105.87312],
"COFS03":	["FS3",	            39.92577,	-105.87337],
"COFS04":	["FS4",	            39.92960,	-105.87340],
"COFS05":	["FS5",	            39.92933,	-105.87081],
"COFS06":	["FS6",	            39.92845,	-105.86850],
"COFS07":	["FS7",	            39.92512,	-105.86994],
"COFS08":	["FS8",	            39.92474,	-105.86826],
"COFS09":	["FS9",	            39.92350,	-105.87027],
"COFS14":	["FS14",	        39.92616,	-105.86673],
"COFS15":	["FS15",	        39.92794,	-105.86460],
"COFS16":	["FS16",	        39.92742,	-105.86334],
"COFS17":	["FS17",	        39.92627,	-105.86912],
"COSBBR":	["Brooklyns",	    37.85724,	-107.72503],
"COSBCR":	["Crux",	        37.90183,	-107.71987],
"COSBME":	["Meadow",	        37.90099,	-107.71702],
"COSBSA":	["Swamp Angel",	    37.90715,	-107.71119],
"COSBSI":	["Silverton",	    37.81433,	-107.66501],
"IDBRB1":	["BB_1",	        43.73154,	-116.13968],
"IDBRB2":	["BB_2",	        43.73302,	-116.12270],
"IDBRB3":	["BB_3",	        43.73636,	-116.12090],
"IDBRB4":	["BB_4",	        43.73747,	-116.12167],
"IDBRB5":	["BB_5",	        43.76006,	-116.10379],
"IDBRB6":	["BB_6",	        43.75717,	-116.09092],
"IDBRB7":	["BB_7",	        43.75729,	-116.09081],
"IDBRL1":	["LR_1",	        44.07249,	-115.51688],
"IDBRL2":	["LR_2",	        44.13563,	-115.31426],
"IDBRL3":	["LR_3",	        44.19866,	-115.24869],
"IDBRL4":	["LR_4",	        44.23075,	-115.21862],
"IDBRL5":	["LR_5",	        44.30344,	-115.23461],
"IDBRL6":	["LR_6",	        44.30309,	-115.23465],
"IDBRL7":	["LR_7",	        44.29675,	-115.23961],
"IDBRL8":	["LR_8",	        44.29088,	-115.24397],
"IDBRL9":	["LR_9",	        44.28330,	-115.24156],
"IDBRM1":	["MC_1",	        43.93208,	-115.66576],
"IDBRM2":	["MC_2",	        43.94678,	-115.66840],
"IDBRM3":	["MC_3",	        43.94743,	-115.67196],
"IDBRM4":	["MC_4",	        43.94786,	-115.67480],
"IDBRM5":	["MC_5",	        43.94737,	-115.67679],
"IDBRM6":	["MC_6",	        43.94854,	-115.67828],
"IDBRM7":	["MC_7",	        43.95258,	-115.67997],
"IDBRM8":	["MC_8",	        43.95249,	-115.68174],
"IDBRM9":	["MC_9",	        43.95490,	-115.68158]
}

# function to rewrite site, lat, and lon
def adjust_intervalboards(filename, coord_dict, nfname):
    boardID = filename.stem.split('_')[0]

    wb = load_workbook(filename)
    ws = wb.active

    ws['D4'].value = coord_dict.get(boardID)[0] # site
    ws['C9'].value = coord_dict.get(boardID)[1] # lat
    ws['E9'].value = coord_dict.get(boardID)[2] # lon

    wb.save(nfname)


for filename in sorted(path_in.rglob('*.xlsx')):
    print(filename.name)
    nfname = outpath.joinpath(filename.name)
    # print(nfname)
    r = adjust_intervalboards(filename, coord_dict, nfname)
